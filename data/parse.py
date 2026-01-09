from openpyxl import load_workbook
import json

wb = load_workbook('Survey Skor UTBK Teknologi Informasi 2025.xlsx')["Form Responses 1"]

scores = []
total_mean = []

for row in wb.iter_rows(min_col=2, max_col=8, min_row=2, values_only=True):
    if row[0] is None:
        break

    mean = 0
    for score in row:
        mean += float(score)
    mean /= len(row)
    total_mean.append(mean)

    scores.append({
        "pu": row[0],
        "ppu": row[1],
        "pbm": row[2],
        "pk": row[3],
        "lbi": row[4],
        "lbe": row[5],
        "pm": row[6],
        "mean": mean
    })

# Sort scores by mean in descending order
scores.sort(key=lambda x: x["mean"], reverse=True)

avg_mean = sum(total_mean) / len(total_mean)
max_mean = max(total_mean)
min_mean = min(total_mean)

# Get min, mean, max of each subject
subject_stats = {}
for subject in ["pu", "ppu", "pbm", "pk", "lbi", "lbe", "pm"]:
    subject_scores = [entry[subject] for entry in scores]
    subject_stats[subject] = {
        "min": min(subject_scores),
        "mean": sum(subject_scores) / len(subject_scores),
        "max": max(subject_scores)
    }

with open('scores.json', 'w') as f:
    json.dump({
        "mean": avg_mean,
        "max": max_mean,
        "min": min_mean,
        "scores": scores,
        "summary": {
            "total_entries": len(scores),
            **subject_stats
        }
    }, f, indent=4)
